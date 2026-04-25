#!/usr/bin/env python3
"""
parse_utils.py: 文档格式解析工具
支持 PDF、Excel、CSV 等格式的内容提取
"""

import os
import json
import tempfile
import subprocess
from pathlib import Path

# 尝试导入可选库
try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


def get_file_token_from_doc(doc_token):
    """根据 doc_token 获取 file_token

    用于 xlsx 等文件类型，需要先获取真实的 file_token 再下载。

    返回: (file_token, error_msg)
    """
    import platform
    # 使用 wiki spaces get_node 获取文件元数据
    # Windows 需要双引号包裹 JSON 并转义，Linux/macOS 用单引号即可
    if platform.system() == 'Windows':
        cmd = f'lark-cli wiki spaces get_node --params "{{\\"token\\":\\"{doc_token}\\"}}"'
    else:
        cmd = f"lark-cli wiki spaces get_node --params '{{\\\"token\\\":\\\"{doc_token}\\\"}}'"
    result = subprocess.run(
        cmd,
        shell=True,
        capture_output=True
    )
    if result.returncode != 0:
        return "", f"获取文件元数据失败: {result.stderr.decode('utf-8', errors='replace')}"

    try:
        data = json.loads(result.stdout.decode('utf-8', errors='replace'))
        if data.get("code") == 0:
            # 从 node 数据中获取 file_token
            node = data.get("data", {}).get("node", {})
            # obj_token 才是真正的 file_token，file_token 字段可能是空的
            file_token = node.get("obj_token", "") or node.get("file_token", "")
            if file_token:
                return file_token, ""
            return "", "未找到 obj_token 或 file_token"
        return "", f"API 错误: {data}"
    except json.JSONDecodeError as e:
        return "", f"解析元数据失败: {e}"


def download_file(file_token, output_filename):
    """使用 lark-cli 下载文件到本地

    Args:
        file_token: 文件 token
        output_filename: 相对文件名（不含路径，lark-cli 要求相对路径）

    返回: (success, error_msg)
    """
    # lark-cli 要求相对路径，不能是 /tmp/xxx
    cmd = f'lark-cli drive +download --file-token "{file_token}" --output "{output_filename}"'
    result = subprocess.run(
        cmd,
        shell=True,
        capture_output=True
    )
    if result.returncode != 0:
        return False, result.stderr.decode('utf-8', errors='replace')
    return True, ""


def cleanup_file(filename):
    """清理下载的文件"""
    try:
        if os.path.exists(filename):
            os.remove(filename)
    except:
        pass


def parse_pdf(file_path, max_chars=500):
    """解析 PDF 文件，返回纯文本

    Args:
        file_path: PDF 文件路径
        max_chars: 最大提取字符数

    Returns:
        (content, error_msg)
    """
    if not HAS_PYPDF2:
        return "", "未安装 PyPDF2，请运行: pip install PyPDF2"

    try:
        text = []
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
                    if sum(len(t) for t in text) > max_chars:
                        break

        content = "\n".join(text)[:max_chars]
        return content, None
    except Exception as e:
        return "", f"PDF 解析失败: {e}"


def parse_excel(file_path, max_chars=500):
    """解析 Excel 文件，返回纯文本

    Args:
        file_path: Excel 文件路径
        max_chars: 最大提取字符数

    Returns:
        (content, error_msg)
    """
    if not HAS_OPENPYXL:
        return "", "未安装 openpyxl，请运行: pip install openpyxl"

    try:
        text_parts = []
        wb = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===")

            for row in ws.iter_rows(values_only=True):
                # 过滤 None 值，拼接行
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)

                if sum(len(t) for t in text_parts) > max_chars:
                    break

        content = "\n".join(text_parts)[:max_chars]
        return content, None
    except Exception as e:
        return "", f"Excel 解析失败: {e}"


def parse_csv(file_path, max_chars=500):
    """解析 CSV 文件，返回纯文本

    Args:
        file_path: CSV 文件路径
        max_chars: 最大提取字符数

    Returns:
        (content, error_msg)
    """
    if not HAS_PANDAS:
        return "", "未安装 pandas，请运行: pip install pandas"

    try:
        df = pd.read_csv(file_path)
        content = df.to_string(max_chars=max_chars)
        return content[:max_chars], None
    except Exception as e:
        return "", f"CSV 解析失败: {e}"


def parse_document(doc_token, doc_type, max_chars=500):
    """统一文档解析入口

    支持类型: PDF, XLSX, XLS, CSV, DOCX, DOC

    Args:
        doc_token: 文档 token（xlsx 等文件类型需要先获取 file_token）
        doc_type: 文档类型
        max_chars: 最大提取字符数

    Returns:
        (content, error_msg)
    """
    if doc_type in ("PDF",):
        # PDF 需要先下载再解析
        output_filename = f"{doc_token}.pdf"
        success, err = download_file(doc_token, output_filename)
        if not success:
            return "", f"下载失败: {err}"
        try:
            content, err = parse_pdf(output_filename, max_chars)
            return content, err
        finally:
            cleanup_file(output_filename)

    elif doc_type in ("XLSX", "XLS"):
        # xlsx 文件需要先获取 file_token
        file_token, err = get_file_token_from_doc(doc_token)
        if err:
            return "", f"获取 file_token 失败: {err}"

        output_filename = f"{file_token}.xlsx"
        success, err = download_file(file_token, output_filename)
        if not success:
            return "", f"下载失败: {err}"
        try:
            content, err = parse_excel(output_filename, max_chars)
            return content, err
        finally:
            cleanup_file(output_filename)

    elif doc_type in ("CSV",):
        # CSV 也需要先获取 file_token
        file_token, err = get_file_token_from_doc(doc_token)
        if err:
            return "", f"获取 file_token 失败: {err}"

        output_filename = f"{file_token}.csv"
        success, err = download_file(file_token, output_filename)
        if not success:
            return "", f"下载失败: {err}"
        try:
            content, err = parse_csv(output_filename, max_chars)
            return content, err
        finally:
            cleanup_file(output_filename)

    else:
        return "", f"不支持的文档类型: {doc_type}"


def main():
    """测试解析功能"""
    print("=== parse_utils.py 测试 ===")
    print(f"PyPDF2: {'已安装' if HAS_PYPDF2 else '未安装'}")
    print(f"openpyxl: {'已安装' if HAS_OPENPYXL else '未安装'}")
    print(f"pandas: {'已安装' if HAS_PANDAS else '未安装'}")
    print()
    print("用法:")
    print("  from parse_utils import parse_document")
    print("  content, err = parse_document(token, 'PDF', max_chars=500)")


if __name__ == "__main__":
    main()
